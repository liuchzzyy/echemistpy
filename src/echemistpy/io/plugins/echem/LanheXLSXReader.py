#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ruff: noqa: N999
"""XLSX Data Reader for LANHE battery test files with metadata extraction using traitlets."""
# ruff: noqa: N999

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import openpyxl.worksheet.worksheet
import pandas as pd
import xarray as xr
from traitlets import HasTraits, List, Unicode

from echemistpy.io.structures import RawData, RawDataInfo

logger = logging.getLogger(__name__)


class LanheXLSXReader(HasTraits):
    """Reader for LANHE exported XLSX files."""

    filepath = Unicode()
    active_material_mass = Unicode(allow_none=True)
    sample_name = Unicode(None, allow_none=True)
    start_time = Unicode(None, allow_none=True)
    instrument = Unicode("LANHE", allow_none=True)
    operator = Unicode(None, allow_none=True)
    wave_number = Unicode(None, allow_none=True)
    technique = List(Unicode(), default_value=["echem"])

    def __init__(self, filepath: str | Path | None = None, **kwargs):
        super().__init__(**kwargs)
        if filepath:
            self.filepath = str(filepath)

    def load(self) -> tuple[RawData, RawDataInfo]:
        """Load LANHE XLSX file(s) and return RawData and RawDataInfo."""
        if not self.filepath:
            raise ValueError("filepath not set")

        path = Path(self.filepath)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        if path.is_file():
            return self._load_single_file(path)
        elif path.is_dir():
            return self._load_directory(path)
        else:
            raise ValueError(f"Path is neither a file nor a directory: {path}")

    def _load_single_file(self, path: Path) -> tuple[RawData, RawDataInfo]:
        """Internal method to load a single LANHE XLSX file."""
        # Read raw data and metadata
        metadata, data_dict = self._read_xlsx(path)

        # Clean metadata and data
        cleaned_metadata = self._clean_metadata(metadata)
        mass = self.active_material_mass or cleaned_metadata.get("active_material_mass")
        cleaned_data = self._clean_data(data_dict, active_material_mass=mass)

        # Convert to xarray Dataset
        ds = xr.Dataset({k: (("record",), v) for k, v in cleaned_data.items()})

        # Handle time coordinates
        if "SysTime" in ds:
            systimes = pd.to_datetime(ds["SysTime"].values)
            ds = ds.assign_coords(systime=(("record",), systimes))
            ds = ds.drop_vars("SysTime")
            # Calculate relative time
            rel_times = systimes - systimes[0]
            ds = ds.assign_coords(time_s=(("record",), rel_times))
            ds.time_s.attrs["units"] = "s"

        # Use Record or similar index as the coordinate for the 'record' dimension
        for index_col in ["Record", "record", "Row", "row", "Index", "index"]:
            if index_col in ds:
                ds = ds.set_index(record=index_col)
                break

        # Create RawData and RawDataInfo
        raw_data = RawData(data=ds)

        # Extract top-level metadata
        start_time_val = self.start_time or cleaned_metadata.get("start_time")
        if isinstance(start_time_val, datetime):
            start_time_val = start_time_val.strftime("%Y-%m-%d %H:%M:%S")

        tech_list = self.technique if self.technique != ["echem"] else [*list(self.technique), "gcd"]

        raw_info = RawDataInfo(
            sample_name=self.sample_name or str(cleaned_metadata.get("test_name", "Unknown")),
            start_time=start_time_val,
            operator=self.operator or cleaned_metadata.get("operator"),
            technique=tech_list,
            instrument=self.instrument,
            active_material_mass=mass,
            wave_number=self.wave_number,
            others={**cleaned_metadata, "file_path": str(path)},
        )

        return raw_data, raw_info

    def _load_directory(self, path: Path) -> tuple[RawData, RawDataInfo]:
        """Load all LANHE XLSX files in a directory and its subdirectories into a DataTree."""
        xlsx_files = sorted(path.rglob("*.xlsx"))
        if not xlsx_files:
            raise FileNotFoundError(f"No .xlsx files found in {path}")

        # Create a root DataTree
        tree = xr.DataTree(name=path.name)
        infos = []

        for f in xlsx_files:
            try:
                raw_data, raw_info = self._load_single_file(f)
                ds = raw_data.data

                # Sanitize variable names for DataTree (replace / with _)
                rename_dict = {str(var): str(var).replace("/", "_") for var in ds.data_vars if "/" in str(var)}
                if rename_dict:
                    ds = ds.rename(rename_dict)

                # Determine relative path for the tree
                rel_path = f.relative_to(path)

                # Build the path string for DataTree (using / as separator)
                node_path = "/".join(rel_path.with_suffix("").parts)

                # Add to tree
                tree[node_path] = ds

                # Store metadata in node attributes
                tree[node_path].attrs.update(raw_info.to_dict())

                infos.append(raw_info)
            except Exception as e:
                logger.warning("Failed to load %s: %s", f, e)

        if not tree.children and not tree.has_data:
            raise RuntimeError(f"Failed to load any .xlsx files from {path}")

        # Merge RawDataInfo for the root
        merged_info = self._merge_infos(infos, path)

        return RawData(data=tree), merged_info

    def _merge_infos(self, infos: list[RawDataInfo], root_path: Path) -> RawDataInfo:
        """Merge multiple RawDataInfo objects into one."""
        if not infos:
            return RawDataInfo()

        # Use the first one as base
        base = infos[0]

        # Collect all techniques
        all_techs = set()
        for info in infos:
            for t in info.technique:
                all_techs.add(t)

        # Create merged info
        merged_info = RawDataInfo(
            sample_name=self.sample_name or root_path.name,
            start_time=self.start_time or base.start_time,
            operator=self.operator or base.operator,
            technique=list(all_techs),
            instrument=self.instrument or base.instrument,
            active_material_mass=self.active_material_mass or base.active_material_mass,
            wave_number=self.wave_number or base.wave_number,
            others={
                "merged_files": [str(info.get("file_path")) for info in infos if info.get("file_path")],
                "n_files": len(infos),
                "structure": "DataTree",
            },
        )

        return merged_info

    def _read_xlsx(self, filepath: Path) -> tuple[dict[str, Any], dict[str, Any]]:
        """Read metadata and data from LANHE .xlsx file."""
        metadata: dict[str, Any] = {}
        data_dict: dict[str, Any] = {}
        data_sheet_name: str | None = None

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            self._read_test_info(wb, metadata)
            self._read_proc_info(wb, metadata)
            self._read_log_info(wb, metadata)
            data_sheet_name = self._find_data_sheet(wb)

            if data_sheet_name:
                data_dict = self._read_record_data_from_ws(wb[data_sheet_name], data_sheet_name)
        finally:
            wb.close()

        return metadata, data_dict

    def _read_record_data_from_ws(self, ws: openpyxl.worksheet.worksheet.Worksheet, data_sheet_name: str) -> dict[str, Any]:
        """Extract record-level data from an open worksheet."""
        header, rows = None, []
        for row in ws.iter_rows(values_only=True):
            if not row[0]:
                continue

            # Header detection: look for "Cycle" and "Step" in the first few columns
            if not header and str(row[0]).strip().lower() == "cycle" and len(row) > 10:
                row_str = " ".join(str(c).lower() for c in row[:10] if c)
                if "step" in row_str:
                    header = [str(c).strip() for c in row if c and str(c).strip()]
                    continue

            # Data extraction: ensure first 3 columns are numeric (Cycle, Step, Index/Record)
            if header:
                try:
                    # LANHE data rows usually start with 3 integers
                    [int(c) for c in row[:3]]
                    rows.append(row)
                except (ValueError, TypeError):
                    continue

        if not (header and rows):
            return {}

        # Transpose rows to columns and convert values
        data = {h: [self._convert_time(r[i]) if i < len(r) else None for r in rows] for i, h in enumerate(header)}
        data["_metadata"] = {"sheet_name": data_sheet_name, "num_rows": len(rows), "columns": header}
        return data

    def _read_test_info(self, wb: openpyxl.Workbook, metadata: dict[str, Any]) -> None:
        """Read Test Information sheet."""
        if "Test information" in wb.sheetnames:
            try:
                ws = wb["Test information"]
                headers = [str(cell.value) for cell in ws[1] if cell.value]
                if ws.max_row >= 2:
                    metadata["Test_Information"] = {h: self._convert_time(ws[2][i].value) for i, h in enumerate(headers) if i < len(ws[2])}
            except Exception as e:
                logger.debug("Error reading Test Information: %s", e)

    def _read_proc_info(self, wb: openpyxl.Workbook, metadata: dict[str, Any]) -> None:
        """Read Ch1_Proc sheet and extract Work Mode table."""
        if "Ch1_Proc" not in wb.sheetnames:
            return
        try:
            ws, proc_info, work_mode, headers = wb["Ch1_Proc"], {}, [], None
            for row in ws.iter_rows(values_only=True):
                if not row[0]:
                    continue
                if str(row[0]) == "Order":
                    headers = [str(c) for c in row if c]
                elif headers:
                    if not str(row[0]).strip():
                        break
                    work_mode.append({h: self._convert_time(row[i]) for i, h in enumerate(headers) if i < len(row)})
                elif row[1]:
                    proc_info[str(row[0])] = self._convert_time(row[1])
            if work_mode:
                proc_info["Work_Mode"] = work_mode
            metadata["Channel_Process_Info"] = proc_info
        except Exception as e:
            logger.debug("Error reading Channel Process Info: %s", e)

    def _read_log_info(self, wb: openpyxl.Workbook, metadata: dict[str, Any]) -> None:
        """Read Log sheet."""
        if "Log" in wb.sheetnames:
            try:
                ws = wb["Log"]
                headers = [str(c.value) for c in ws[1] if c.value]
                metadata["Log_Info"] = [{h: self._convert_time(r[i]) for i, h in enumerate(headers) if i < len(r)} for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
            except Exception as e:
                logger.debug("Error reading Log Info: %s", e)

    @staticmethod
    def _find_data_sheet(wb: openpyxl.Workbook) -> str | None:
        """Find DefaultGroup sheet."""
        for sheet_name in wb.sheetnames:
            if "DefaultGroup" in sheet_name:
                return sheet_name
        return None

    @staticmethod
    def _convert_time(value: Any) -> Any:
        """Convert Excel values (datetime, timedelta, or LANHE time strings) to standard formats."""
        if value is None or isinstance(value, datetime):
            return value

        if hasattr(value, "total_seconds"):
            return value.total_seconds()

        if not (isinstance(value, str) and ":" in value):
            return value

        if " " in value:
            parts = value.split(" ", 1)
            # Try absolute time first, then duration
            return LanheXLSXReader._parse_abs_time(parts[0], parts[1]) or LanheXLSXReader._parse_duration(parts[0], parts[1]) or value

        if len(value) == 10 and value[4] in {"-", "/"}:
            try:
                return datetime(int(value[0:4]), int(value[5:7]), int(value[8:10]))
            except (ValueError, IndexError):
                pass
        return value

    @staticmethod
    def _parse_abs_time(first: str, second: str) -> datetime | None:
        """Parse YYYY-MM-DD HH:MM:SS.mmm format."""
        if len(first) == 10 and first[4] in {"-", "/"}:
            try:
                year, month, day = int(first[0:4]), int(first[5:7]), int(first[8:10])
                hour, minute = int(second[0:2]), int(second[3:5])
                if "." in second:
                    sec_parts = second[6:].split(".")
                    usec = int(sec_parts[1].ljust(6, "0")[:6])
                    return datetime(year, month, day, hour, minute, int(sec_parts[0]), usec)
                return datetime(year, month, day, hour, minute, int(second[6:8]))
            except (ValueError, IndexError):
                pass
        return None

    @staticmethod
    def _parse_duration(first: str, second: str) -> float | None:
        """Parse D HH:MM:SS.mmm format."""
        if len(first) <= 3 and first.isdigit():
            try:
                hms = second.split(":")
                if len(hms) == 3:
                    return int(first) * 86400 + int(hms[0]) * 3600 + int(hms[1]) * 60 + float(hms[2])
            except (ValueError, IndexError):
                pass
        return None

    @staticmethod
    def _clean_metadata(metadata: dict[str, Any]) -> dict[str, Any]:
        """Clean metadata to keep only essential fields."""
        cleaned = {}

        # Test Information
        if info := metadata.get("Test_Information"):
            mapping = {
                "Test name": "test_name",
                "Start time": "start_time",
                "Finish time": "finish_time",
                "Active material": "active_material",
                "Operator": "operator",
            }
            for raw_key, clean_key in mapping.items():
                # Case-insensitive lookup
                actual_key = next((k for k in info if k.lower() == raw_key.lower()), None)
                if actual_key:
                    cleaned[clean_key] = info[actual_key]

            # Parse Active material string: "Nominal specific capacity: 372 mAh/g Active material: 1.23 mg"
            if am := cleaned.get("active_material"):
                am_str = str(am)
                if "Active material:" in am_str:
                    parts = am_str.split("Active material:")
                    cleaned["active_material_mass"] = parts[1].strip()
                    if "Nominal specific capacity:" in parts[0]:
                        cleaned["nominal_specific_capacity"] = parts[0].replace("Nominal specific capacity:", "").strip()

        # Channel Process Info
        if proc := metadata.get("Channel_Process_Info"):
            proc_fields = ["Channel Number", "Name", "Description", "Unit Scheme", "Safety", "Work_Mode"]
            cleaned["channel_process_info"] = {k.lower().replace(" ", "_"): proc[k] for k in proc_fields if k in proc}

        # Ensure technique is a list
        tech = metadata.get("technique", ["echem"])
        cleaned["technique"] = [tech] if isinstance(tech, str) else tech

        return {**metadata, **cleaned}

    @staticmethod
    def _clean_data(data: dict[str, Any], active_material_mass: Any = None) -> dict[str, Any]:
        """Return data with original headers, filtering for requested columns and ordering them."""
        # 定义固定的列名顺序
        ordered_cols = [
            "Record",
            "SysTime",
            "Cycle",
            "TestTime",
            "Voltage/V",
            "Current/uA",
            "Capacity/uAh",
            "SpeCap/mAh/g",
            "SpeCap_cal/mAh/g",
            "dQdV/uAh/V",
            "dVdQ/V/uAh",
        ]

        # 1. 预计算 SpeCap_cal/mAh/g (如果提供了质量)
        spe_cap_cal = None
        if active_material_mass and "Capacity/uAh" in data:
            try:
                mass_str = str(active_material_mass).lower()
                # 提取数值
                mass_val = float("".join(c for c in mass_str if c.isdigit() or c == "."))
                # 确定单位系数
                factor = 0.001 if "mg" in mass_str else 1.0
                mass_g = mass_val * factor

                if mass_g > 0:
                    # SpeCap (mAh/g) = (Capacity (uAh) / 1000) / mass (g)
                    spe_cap_cal = [(float(c) / 1000.0) / mass_g if c is not None else None for c in data["Capacity/uAh"]]
            except (ValueError, TypeError, ZeroDivisionError):
                logger.debug("Failed to calculate SpeCap_cal/mAh/g with mass: %s", active_material_mass)

        # 2. 按照固定顺序构建结果字典
        cleaned_data = {}
        for col in ordered_cols:
            if col == "SpeCap_cal/mAh/g":
                if spe_cap_cal is not None:
                    cleaned_data[col] = spe_cap_cal
            elif col in data:
                cleaned_data[col] = data[col]

        return cleaned_data
