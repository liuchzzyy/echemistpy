#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ruff: noqa: N999
"""XLSX Data Reader for LANHE battery test files with metadata extraction using traitlets."""

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, ClassVar, Dict, List, Optional, Tuple, Union

import openpyxl
import openpyxl.worksheet.worksheet
import pandas as pd
import xarray as xr
from traitlets import HasTraits, Unicode
from traitlets import List as TList

from echemistpy.io.structures import RawData, RawDataInfo

logger = logging.getLogger(__name__)


class LanheXLSXReader(HasTraits):
    """Reader for LANHE exported XLSX files.

    This reader handles the specific structure of LANHE battery test exports,
    including metadata from multiple sheets and time-series data from the main data sheet.
    """

    # --- Constants for better maintainability ---
    ORDERED_COLUMNS: ClassVar[List[str]] = [
        "Record",
        "SysTime",
        "Cycle",
        "TestTime",
        "Voltage/V",
        "Current/uA",
        "Capacity/uAh",
        "SpeCap/mAh/g",
        "SpeCap_cal/mAh/g",
        "dQdV/uAh/V",
        "dVdQ/V/uAh",
    ]

    METADATA_MAPPING: ClassVar[Dict[str, str]] = {
        "Test name": "test_name",
        "Start time": "start_time",
        "Finish time": "finish_time",
        "Active material": "active_material",
        "Operator": "operator",
    }

    INDEX_COLUMNS: ClassVar[List[str]] = ["Record", "record", "Row", "row", "Index", "index"]

    # --- Loader Metadata ---
    supports_directories: ClassVar[bool] = True
    instrument: ClassVar[str] = "lanhe"

    # --- Traitlets for configuration and metadata ---
    filepath = Unicode()
    active_material_mass = Unicode(allow_none=True)
    sample_name = Unicode(None, allow_none=True)
    start_time = Unicode(None, allow_none=True)
    operator = Unicode(None, allow_none=True)
    wave_number = Unicode(None, allow_none=True)
    technique = TList(Unicode(), default_value=["echem"])
    # instrument traitlet removed to avoid conflict with ClassVar

    def __init__(self, filepath: Optional[Union[str, Path]] = None, **kwargs: Any) -> None:
        super().__init__(**kwargs)
        if filepath:
            self.filepath = str(filepath)

    def load(self, **_kwargs: Any) -> Tuple[RawData, RawDataInfo]:
        """Load LANHE XLSX file(s) and return RawData and RawDataInfo.

        Returns:
            A tuple containing the loaded RawData and its associated RawDataInfo.

        Raises:
            ValueError: If filepath is not set or path type is invalid.
            FileNotFoundError: If the specified path does not exist.
        """
        if not self.filepath:
            raise ValueError("filepath must be set before calling load()")

        path = Path(self.filepath)
        if not path.exists():
            raise FileNotFoundError(f"Path not found: {path}")

        if path.is_file():
            return self._load_single_file(path)
        if path.is_dir():
            return self._load_directory(path)

        raise ValueError(f"Path is neither a file nor a directory: {path}")

    def _load_single_file(self, path: Path) -> Tuple[RawData, RawDataInfo]:
        """Internal method to load and process a single LANHE XLSX file."""
        # 1. Extraction
        metadata, data_dict = self._read_xlsx(path)

        # 2. Cleaning & Standardization
        cleaned_metadata = self._clean_metadata(metadata)
        mass = self.active_material_mass or cleaned_metadata.get("active_material_mass")
        cleaned_data = self._clean_data(data_dict, active_material_mass=mass)

        # 3. Xarray Conversion
        ds = self._create_dataset(cleaned_data)
        ds = self._apply_time_coords(ds)
        ds = self._set_primary_index(ds)

        # 4. Metadata Packaging
        raw_info = self._create_raw_info(path, cleaned_metadata, mass)

        return RawData(data=ds), raw_info

    @staticmethod
    def _create_dataset(data: Dict[str, Any]) -> xr.Dataset:
        """Create an xarray Dataset from cleaned data dictionary."""
        ds = xr.Dataset({k: (("record",), v) for k, v in data.items()})

        # Sanitize variable names (replace / with _)
        rename_dict = {str(var): str(var).replace("/", "_") for var in ds.data_vars if "/" in str(var)}
        if rename_dict:
            ds = ds.rename(rename_dict)

        return ds

    @staticmethod
    def _apply_time_coords(ds: xr.Dataset) -> xr.Dataset:
        """Convert SysTime to coordinates and calculate relative time."""
        systime_key = "SysTime" if "SysTime" in ds else "SysTime".replace("/", "_")

        if systime_key in ds:
            try:
                systimes = pd.to_datetime(ds[systime_key].values)
                ds = ds.assign_coords(systime=(("record",), systimes))
                ds = ds.drop_vars(systime_key)

                # Calculate relative time in seconds
                rel_times = (systimes - systimes[0]).total_seconds()
                ds = ds.assign_coords(time_s=(("record",), rel_times))
                ds.time_s.attrs.update({"units": "s", "long_name": "Relative Time"})
                ds.systime.attrs.update({"long_name": "System Time"})
            except Exception as e:
                logger.warning("Failed to process time coordinates: %s", e)

        return ds

    @staticmethod
    def _set_primary_index(ds: xr.Dataset) -> xr.Dataset:
        """Set the primary index for the 'record' dimension."""
        for index_col in LanheXLSXReader.INDEX_COLUMNS:
            # Check both original and sanitized names
            for col in [index_col, index_col.replace("/", "_")]:
                if col in ds:
                    return ds.set_index(record=col)
        return ds

    def _create_raw_info(self, path: Path, metadata: Dict[str, Any], mass: Any) -> RawDataInfo:
        """Create a RawDataInfo object from metadata."""
        start_time_val = self.start_time or metadata.get("start_time")
        if isinstance(start_time_val, datetime):
            start_time_val = start_time_val.strftime("%Y-%m-%d %H:%M:%S")

        # Default to GCD if it's echem
        tech_list = list(self.technique)
        if tech_list == ["echem"]:
            tech_list.append("gcd")

        return RawDataInfo(
            sample_name=self.sample_name or str(metadata.get("test_name", path.stem)),
            start_time=start_time_val,
            operator=self.operator or metadata.get("operator"),
            technique=tech_list,
            instrument=self.instrument,
            active_material_mass=mass,
            wave_number=self.wave_number,
            others={**metadata, "file_path": str(path)},
        )

    def _load_directory(self, path: Path) -> Tuple[RawData, RawDataInfo]:
        """Load all LANHE XLSX files in a directory into a DataTree."""
        xlsx_files = sorted(path.rglob("*.xlsx"))
        if not xlsx_files:
            raise FileNotFoundError(f"No .xlsx files found in {path}")

        tree = xr.DataTree(name=path.name)
        infos = []

        for f in xlsx_files:
            try:
                raw_data, raw_info = self._load_single_file(f)
                node_path = "/".join(f.relative_to(path).with_suffix("").parts)

                tree[node_path] = raw_data.data
                tree[node_path].attrs.update(raw_info.to_dict())
                infos.append(raw_info)
            except Exception as e:
                logger.error("Failed to load %s: %s", f, e)

        if not tree.children and not tree.has_data:
            raise RuntimeError(f"Failed to load any valid .xlsx files from {path}")

        return RawData(data=tree), self._merge_infos(infos, path)

    def _merge_infos(self, infos: List[RawDataInfo], root_path: Path) -> RawDataInfo:
        """Merge multiple RawDataInfo objects into a single root info."""
        if not infos:
            return RawDataInfo()

        base = infos[0]
        all_techs = {t for info in infos for t in info.technique}

        return RawDataInfo(
            sample_name=self.sample_name or root_path.name,
            start_time=self.start_time or base.start_time,
            operator=self.operator or base.operator,
            technique=list(all_techs),
            instrument=self.instrument or base.instrument,
            active_material_mass=self.active_material_mass or base.active_material_mass,
            wave_number=self.wave_number or base.wave_number,
            others={
                "merged_files": [str(info.get("file_path")) for info in infos if info.get("file_path")],
                "n_files": len(infos),
                "structure": "DataTree",
            },
        )

    def _read_xlsx(self, filepath: Path) -> Tuple[Dict[str, Any], Dict[str, Any]]:
        """Read metadata and data from LANHE .xlsx file using openpyxl."""
        metadata: Dict[str, Any] = {}
        data_dict: Dict[str, Any] = {}

        # Use read_only for memory efficiency
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            self._read_test_info(wb, metadata)
            self._read_proc_info(wb, metadata)
            self._read_log_info(wb, metadata)

            if data_sheet_name := self._find_data_sheet(wb):
                data_dict = self._read_record_data_from_ws(wb[data_sheet_name], data_sheet_name)
            else:
                logger.warning("No data sheet (DefaultGroup) found in %s", filepath)
        finally:
            wb.close()

        return metadata, data_dict

    def _read_record_data_from_ws(self, ws: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str) -> Dict[str, Any]:
        """Extract record-level data from an open worksheet."""
        header: Optional[List[str]] = None
        rows: List[Tuple[Any, ...]] = []

        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue

            # 1. Header Detection
            if not header:
                row_str = " ".join(str(c).lower() for c in row[:10] if c)
                if "cycle" in row_str and "step" in row_str:
                    header = [str(c).strip() for c in row if c and str(c).strip()]
                continue

            # 2. Data Row Validation (First 3 columns should be numeric)
            try:
                [int(c) for c in row[:3]]
                rows.append(row)
            except (ValueError, TypeError):
                continue

        if not (header and rows):
            return {}

        # 3. Data Conversion
        data = {h: [self._convert_time(r[i]) if i < len(r) else None for r in rows] for i, h in enumerate(header)}
        data["_metadata"] = {"sheet_name": sheet_name, "num_rows": len(rows), "columns": header}
        return data

    def _read_test_info(self, wb: openpyxl.Workbook, metadata: Dict[str, Any]) -> None:
        """Read 'Test information' sheet."""
        if "Test information" in wb.sheetnames:
            try:
                ws = wb["Test information"]
                headers = [str(cell.value) for cell in ws[1] if cell.value]
                if ws.max_row >= 2:
                    metadata["Test_Information"] = {h: self._convert_time(ws[2][i].value) for i, h in enumerate(headers) if i < len(ws[2])}
            except Exception as e:
                logger.debug("Error reading Test Information: %s", e)

    def _read_proc_info(self, wb: openpyxl.Workbook, metadata: Dict[str, Any]) -> None:
        """Read 'Ch1_Proc' sheet and extract Work Mode table."""
        if "Ch1_Proc" not in wb.sheetnames:
            return
        try:
            ws = wb["Ch1_Proc"]
            proc_info: Dict[str, Any] = {}
            work_mode: List[Dict[str, Any]] = []
            headers: Optional[List[str]] = None

            for row in ws.iter_rows(values_only=True):
                if not row[0]:
                    continue
                if str(row[0]) == "Order":
                    headers = [str(c) for c in row if c]
                elif headers:
                    if not str(row[0]).strip():
                        break
                    work_mode.append({h: self._convert_time(row[i]) for i, h in enumerate(headers) if i < len(row)})
                elif row[1]:
                    proc_info[str(row[0])] = self._convert_time(row[1])

            if work_mode:
                proc_info["Work_Mode"] = work_mode
            metadata["Channel_Process_Info"] = proc_info
        except Exception as e:
            logger.debug("Error reading Channel Process Info: %s", e)

    def _read_log_info(self, wb: openpyxl.Workbook, metadata: Dict[str, Any]) -> None:
        """Read 'Log' sheet."""
        if "Log" in wb.sheetnames:
            try:
                ws = wb["Log"]
                headers = [str(c.value) for c in ws[1] if c.value]
                metadata["Log_Info"] = [{h: self._convert_time(r[i]) for i, h in enumerate(headers) if i < len(r)} for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
            except Exception as e:
                logger.debug("Error reading Log Info: %s", e)

    @staticmethod
    def _find_data_sheet(wb: openpyxl.Workbook) -> Optional[str]:
        """Find the sheet containing 'DefaultGroup'."""
        return next((name for name in wb.sheetnames if "DefaultGroup" in name), None)

    @staticmethod
    def _convert_time(value: Any) -> Any:
        """Convert Excel values or LANHE time strings to standard formats."""
        if value is None or isinstance(value, (datetime, int, float)):
            return value

        if hasattr(value, "total_seconds"):
            return value.total_seconds()

        if not isinstance(value, str):
            return value

        # Handle LANHE specific time strings
        if ":" in value:
            if " " in value:
                parts = value.split(" ", 1)
                return LanheXLSXReader._parse_abs_time(parts[0], parts[1]) or LanheXLSXReader._parse_duration(parts[0], parts[1]) or value

            # Simple date YYYY-MM-DD
            if len(value) == 10 and value[4] in {"-", "/"}:
                try:
                    return datetime.strptime(value.replace("/", "-"), "%Y-%m-%d")
                except ValueError:
                    pass
        return value

    @staticmethod
    def _parse_abs_time(date_part: str, time_part: str) -> Optional[datetime]:
        """Parse YYYY-MM-DD HH:MM:SS.mmm format."""
        try:
            # Normalize separator
            date_str = date_part.replace("/", "-")
            # Handle milliseconds if present
            if "." in time_part:
                return datetime.strptime(f"{date_str} {time_part}", "%Y-%m-%d %H:%M:%S.%f")
            return datetime.strptime(f"{date_str} {time_part}", "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return None

    @staticmethod
    def _parse_duration(days_part: str, hms_part: str) -> Optional[float]:
        """Parse D HH:MM:SS.mmm format into total seconds."""
        if days_part.isdigit():
            try:
                hms = hms_part.split(":")
                if len(hms) == 3:
                    return int(days_part) * 86400 + int(hms[0]) * 3600 + int(hms[1]) * 60 + float(hms[2])
            except (ValueError, IndexError):
                pass
        return None

    def _clean_metadata(self, metadata: Dict[str, Any]) -> Dict[str, Any]:
        """Clean and map metadata to standard fields."""
        cleaned: Dict[str, Any] = {}

        if info := metadata.get("Test_Information"):
            for raw_key, clean_key in self.METADATA_MAPPING.items():
                actual_key = next((k for k in info if k.lower() == raw_key.lower()), None)
                if actual_key:
                    cleaned[clean_key] = info[actual_key]

            # Parse Active material mass and capacity
            if am := cleaned.get("active_material"):
                am_str = str(am)
                if "Active material:" in am_str:
                    parts = am_str.split("Active material:")
                    cleaned["active_material_mass"] = parts[1].strip()
                    if "Nominal specific capacity:" in parts[0]:
                        cleaned["nominal_specific_capacity"] = parts[0].replace("Nominal specific capacity:", "").strip()

        if proc := metadata.get("Channel_Process_Info"):
            proc_fields = ["Channel Number", "Name", "Description", "Unit Scheme", "Safety", "Work_Mode"]
            cleaned["channel_process_info"] = {k.lower().replace(" ", "_"): proc[k] for k in proc_fields if k in proc}

        cleaned["technique"] = list(self.technique)
        return {**metadata, **cleaned}

    def _clean_data(self, data: Dict[str, Any], active_material_mass: Any = None) -> Dict[str, Any]:
        """Filter, order, and optionally calculate specific capacity."""
        # 1. Calculate Specific Capacity if mass is available
        spe_cap_cal = self._calculate_specific_capacity(data, active_material_mass)

        # 2. Build ordered result dictionary
        cleaned_data = {}
        for col in self.ORDERED_COLUMNS:
            if col == "SpeCap_cal/mAh/g":
                if spe_cap_cal is not None:
                    cleaned_data[col] = spe_cap_cal
            elif col in data:
                cleaned_data[col] = data[col]

        return cleaned_data

    @staticmethod
    def _calculate_specific_capacity(data: Dict[str, Any], mass_input: Any) -> Optional[List[Optional[float]]]:
        """Calculate specific capacity (mAh/g) from capacity (uAh) and mass."""
        if not mass_input or "Capacity/uAh" not in data:
            return None

        try:
            mass_str = str(mass_input).lower()
            # Extract numeric value
            match = re.search(r"(\d+\.?\d*)", mass_str)
            if not match:
                return None

            mass_val = float(match.group(1))
            # Unit conversion (default to mg if not specified, but check for g)
            factor = 1.0 if " g" in mass_str or (mass_str.endswith("g") and not mass_str.endswith("mg")) else 0.001
            mass_g = mass_val * factor

            if mass_g > 0:
                # (uAh / 1000) -> mAh; mAh / g -> mAh/g
                return [(float(c) / 1000.0) / mass_g if c is not None else None for c in data["Capacity/uAh"]]
        except (ValueError, TypeError, ZeroDivisionError):
            logger.debug("Failed to calculate specific capacity with mass: %s", mass_input)

        return None
