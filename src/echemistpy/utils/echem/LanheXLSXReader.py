#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""XLSX Data Reader for LANHE battery test files with metadata extraction using traitlets."""

import argparse
import json
import logging
import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from traitlets import HasTraits, Union, Dict, Instance, Unicode, validate

logger = logging.getLogger(__name__)


class XLSXDataReader(HasTraits):
    """Reader for LANHE exported XLSX files with traitlets support."""

    # Traitlets properties
    filepath = Union([Instance(Path), Unicode()], allow_none=True, help="Path to LANHE XLSX file or folder").tag(config=True)

    metadata = Dict(help="Extracted metadata from XLSX file")
    data = Dict(help="Measurement data from XLSX file")

    def __init__(self, filepath: Path | str | None = None, **kwargs):
        """Initialize reader with optional filepath using traitlets."""
        super().__init__(**kwargs)
        if filepath is not None:
            self.filepath = filepath  # Validator will handle conversion

    @validate("filepath")
    def _validate_filepath(self, proposal):
        """Validate filepath and convert to Path if needed."""
        value = proposal["value"]
        if value is None or isinstance(value, Path):
            return value
        if isinstance(value, str):
            return Path(value)
        raise TypeError(f"filepath must be Path or str, got {type(value)}")

    def _read_xlsx(self) -> tuple[dict[str, Any], dict[str, Any]]:
        """Read metadata and data from XLSX file."""
        metadata: dict[str, Any] = {}
        data_dict: dict[str, Any] = {}
        data_sheet_name: str | None = None

        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)

        try:
            # Read Test Information sheet
            if "Test information" in wb.sheetnames:
                try:
                    ws = wb["Test information"]
                    headers = [str(cell.value) for cell in ws[1] if cell.value]
                    if ws.max_row >= 2:
                        test_info = {}
                        for i, header in enumerate(headers):
                            if i < len(ws[2]):
                                value = self._convert_value(ws[2][i].value)
                                test_info[header] = value
                        metadata["Test_Information"] = test_info
                except Exception as e:
                    logger.debug(f"Error reading Test Information: {e}")

            # Read Ch1_Proc sheet and extract Work Mode table
            if "Ch1_Proc" in wb.sheetnames:
                try:
                    ws = wb["Ch1_Proc"]
                    proc_info = {}
                    work_mode_table = []
                    table_headers = None

                    for row in ws.iter_rows(values_only=True):
                        if not row[0]:
                            continue
                        first_col = str(row[0]) if row[0] else ""
                        if first_col == "Order":
                            table_headers = [str(cell) if cell else "" for cell in row]
                            table_headers = [h for h in table_headers if h]
                            continue
                        if table_headers:
                            if not row[0] or (isinstance(row[0], str) and row[0].strip() == ""):
                                break
                            table_row = {}
                            for i, header in enumerate(table_headers):
                                value = self._convert_value(row[i]) if i < len(row) else None
                                table_row[header] = value
                            work_mode_table.append(table_row)
                        elif row[1]:
                            proc_info[first_col] = self._convert_value(row[1])
                    if work_mode_table:
                        proc_info["Work_Mode"] = work_mode_table
                    metadata["Channel_Process_Info"] = proc_info
                except Exception as e:
                    logger.debug(f"Error reading Channel Process Info: {e}")

            # Read Log sheet
            if "Log" in wb.sheetnames:
                try:
                    ws = wb["Log"]
                    headers = [str(cell.value) for cell in ws[1] if cell.value]
                    log_info: list[dict[str, Any]] = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            log_entry = {}
                            for i, header in enumerate(headers):
                                value = self._convert_value(row[i]) if i < len(row) else None
                                log_entry[header] = value
                            log_info.append(log_entry)
                    metadata["Log_Info"] = log_info
                except Exception as e:
                    logger.debug(f"Error reading Log Info: {e}")

            # Find DefaultGroup sheet
            for sheet_name in wb.sheetnames:
                if "DefaultGroup" in sheet_name:
                    data_sheet_name = sheet_name
                    break

        finally:
            wb.close()

        # Normalize metadata keys
        metadata = XLSXDataReader._normalize_dict_keys(metadata)

        # Re-open workbook to extract record-level data
        if data_sheet_name:
            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            try:
                ws = wb[data_sheet_name]

                # Find the "Cycle, Step, Record" header row and extract data from there
                record_header_row = None
                record_data_rows = []

                for row in ws.iter_rows(values_only=True):
                    if not row[0]:
                        continue

                    # Look for the record-level header (starts with "Cycle")
                    if str(row[0]).strip().lower() == "cycle" and len(row) > 15:
                        # Check if it has the characteristic record columns
                        row_str = " ".join(str(cell).lower() for cell in row[:10] if cell)
                        if "step" in row_str and "record" in row_str:
                            record_header_row = [str(cell) if cell else "" for cell in row]
                            record_header_row = [h.strip() for h in record_header_row if h.strip()]
                            continue

                    # Extract record data rows (numeric cycle + step + record)
                    if record_header_row and row[0] and row[1] and row[2]:
                        try:
                            int(row[0])  # cycle
                            int(row[1])  # step
                            int(row[2])  # record
                            workmode = str(row[3]) if len(row) > 3 else ""
                            if any(kw in workmode for kw in ["REST", "CRATE", "LOOP", "END", "CHARGE", "DISCHARGE"]):
                                record_data_rows.append(row)
                        except (ValueError, TypeError):
                            pass

                # Convert record data to dict
                if record_header_row and record_data_rows:
                    # Create dict from extracted data
                    data_dict = {}
                    for col_idx, header in enumerate(record_header_row):
                        normalized_header = XLSXDataReader._normalize_key(header)
                        col_data = []
                        for row_data in record_data_rows:
                            if col_idx < len(row_data):
                                col_data.append(XLSXDataReader._convert_value(row_data[col_idx]))
                            else:
                                col_data.append(None)
                        data_dict[normalized_header] = col_data

                    data_dict["_metadata"] = {
                        "sheet_name": data_sheet_name,
                        "num_rows": len(record_data_rows),
                        "num_columns": len(record_header_row),
                        "columns": [XLSXDataReader._normalize_key(h) for h in record_header_row],
                    }
            finally:
                wb.close()

        return metadata, data_dict

    @staticmethod
    def _convert_value(value: Any) -> Any:
        """Convert datetime/timedelta objects to string format."""
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        elif hasattr(value, "total_seconds"):
            return str(value)
        return value

    @staticmethod
    def _normalize_key(key: str) -> str:
        """Normalize key names: lowercase, spaces/slashes/parentheses to underscores."""
        key = str(key).lower()
        key = key.replace(" ", "_").replace("/", "_")
        key = key.replace("(", "").replace(")", "")
        key = re.sub(r"_+", "_", key)
        return key.strip("_")

    @staticmethod
    def _normalize_dict_keys(data: dict) -> dict:
        """Recursively normalize all dictionary keys."""
        if not isinstance(data, dict):
            return data
        normalized: dict[str, Any] = {}
        for key, value in data.items():
            new_key = XLSXDataReader._normalize_key(key)
            if isinstance(value, dict):
                normalized[new_key] = XLSXDataReader._normalize_dict_keys(value)
            elif isinstance(value, list):
                normalized[new_key] = [XLSXDataReader._normalize_dict_keys(item) if isinstance(item, dict) else item for item in value]
            else:
                normalized[new_key] = value
        return normalized

    @staticmethod
    def _clean_metadata(metadata: dict[str, Any]) -> dict[str, Any]:
        """Clean metadata to keep only essential fields."""
        cleaned = {}

        # Test Information fields to keep (already normalized)
        test_info_keys = ["test_name", "start_time", "finish_time", "active_material"]
        if "test_information" in metadata:
            test_info = metadata["test_information"]
            cleaned_test_info = {k: test_info[k] for k in test_info_keys if k in test_info}

            # Parse and split "active_material" field if needed
            if "active_material" in cleaned_test_info:
                active_material_str = str(cleaned_test_info["active_material"])
                # Parse: "Nominal specific capacity: 308 mAh/g Active material: 1 mg"
                if " Active material: " in active_material_str:
                    parts = active_material_str.split(" Active material: ")
                    if len(parts) == 2:
                        # Extract capacity (remove "Nominal specific capacity: " prefix)
                        capacity = parts[0].replace("Nominal specific capacity: ", "").strip()
                        material = parts[1].strip()
                        cleaned_test_info["nominal_specific_capacity"] = capacity
                        cleaned_test_info["active_material"] = material

            cleaned["test_information"] = cleaned_test_info

        # Channel Process Info fields to keep (already normalized)
        proc_info_keys = ["channel_number", "name", "description", "unit_scheme", "safety", "work_mode"]
        if "channel_process_info" in metadata:
            proc_info = metadata["channel_process_info"]
            cleaned["channel_process_info"] = {k: proc_info[k] for k in proc_info_keys if k in proc_info}

        return cleaned

    @staticmethod
    def _clean_data(data: dict[str, Any]) -> pd.DataFrame:
        """Clean data to keep only specified columns from record-level data."""
        columns_to_keep = [
            "record",
            "systime",
            "cycle",
            "voltage_v",
            "current_ua",
            "capacity_uah",
            "specap_mah_g",
            "dqdv_uah_v",
            "dvdq_v_uah",
        ]

        data_copy = {k: v for k, v in data.items() if k != "_metadata"}
        df = pd.DataFrame(data_copy)
        available_columns = [col for col in columns_to_keep if col in df.columns]
        if available_columns:
            return df[available_columns]
        return pd.DataFrame({col: [] for col in columns_to_keep})

    @staticmethod
    def _save_data(data: dict[str, Any] | list[Any] | pd.DataFrame, output_path: Path, data_type: str, cleaned: bool = False) -> None:
        """Save data: metadata as JSON, data as CSV."""
        try:
            if cleaned:
                output_path = output_path.with_stem(output_path.stem + "_cleaned")
            if data_type == "metadata":
                with open(output_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
            else:
                if isinstance(data, pd.DataFrame):
                    df = data
                else:
                    assert isinstance(data, dict), "data must be dict or DataFrame"
                    data_copy = {k: v for k, v in data.items() if k != "_metadata"}
                    if not data_copy:
                        return
                    df = pd.DataFrame(data_copy)
                df.to_csv(output_path.with_suffix(".csv"), index=False, encoding="utf-8-sig")
        except IOError as e:
            logger.error(f"Failed to save {data_type} to {output_path}: {e}")

    @staticmethod
    def _read_all_xlsx(folder_path: Path | str | None = None) -> dict[str, tuple[dict[str, Any], dict[str, Any]]]:
        """Read all XLSX files from folder."""
        if folder_path is None:
            return {}

        folder_path = Path(folder_path) if isinstance(folder_path, str) else folder_path
        reader = XLSXDataReader()
        all_data = {}

        for xlsx_file in folder_path.glob("*.xlsx"):
            try:
                reader.filepath = xlsx_file
                metadata, data = reader._read_xlsx()
                all_data[xlsx_file.stem] = (metadata, data)
            except Exception as e:
                logger.error(f"Error processing {xlsx_file.name}: {e}", exc_info=True)

        return all_data

    @staticmethod
    def load(filepath: Path | str) -> tuple[dict[str, Any], dict[str, Any]] | dict[str, tuple[dict[str, Any], dict[str, Any]]]:
        """Load XLSX file or folder and return data."""
        filepath = Path(filepath) if isinstance(filepath, str) else filepath
        if filepath.is_dir():
            return XLSXDataReader._read_all_xlsx(filepath)
        if filepath.is_file():
            reader = XLSXDataReader(filepath)
            return reader._read_xlsx()
        raise FileNotFoundError(f"Path not found: {filepath}")

    @staticmethod
    def save(input_filepath: Path | str, output_dir: Path | str, save_cleaned: bool = True, save_original: bool = False) -> None:
        """Load from input and save metadata (JSON) and data (CSV).

        Parameters
        ----------
        input_filepath : Path or str
            Input XLSX file or folder path
        output_dir : Path or str
            Output directory
        save_cleaned : bool
            If True, save cleaned versions (default: True)
        save_original : bool
            If True, also save original (non-cleaned) versions (default: False)
        """
        input_filepath = Path(input_filepath) if isinstance(input_filepath, str) else input_filepath
        output_dir = Path(output_dir) if isinstance(output_dir, str) else output_dir
        output_dir.mkdir(parents=True, exist_ok=True)

        result = XLSXDataReader.load(input_filepath)

        if input_filepath.is_file():
            assert isinstance(result, tuple), "Expected tuple for file"
            metadata, data = result
            metadata_file = output_dir.joinpath(f"{input_filepath.stem}_metadata.json")
            data_file = output_dir.joinpath(f"{input_filepath.stem}_data.json")

            if save_original:
                XLSXDataReader._save_data(metadata, metadata_file, "metadata", cleaned=False)
                XLSXDataReader._save_data(data, data_file, "data", cleaned=False)

            if save_cleaned:
                cleaned_metadata = XLSXDataReader._clean_metadata(metadata)
                cleaned_data = XLSXDataReader._clean_data(data)
                XLSXDataReader._save_data(cleaned_metadata, metadata_file, "metadata", cleaned=True)
                XLSXDataReader._save_data(cleaned_data, data_file, "data", cleaned=True)
        else:
            assert isinstance(result, dict), "Expected dict for folder"
            for filename_stem, (metadata, data) in result.items():
                metadata_file = output_dir.joinpath(f"{filename_stem}_metadata.json")
                data_file = output_dir.joinpath(f"{filename_stem}_data.json")

                if save_original:
                    XLSXDataReader._save_data(metadata, metadata_file, "metadata", cleaned=False)
                    XLSXDataReader._save_data(data, data_file, "data", cleaned=False)

                if save_cleaned:
                    cleaned_metadata = XLSXDataReader._clean_metadata(metadata)
                    cleaned_data = XLSXDataReader._clean_data(data)
                    XLSXDataReader._save_data(cleaned_metadata, metadata_file, "metadata", cleaned=True)
                    XLSXDataReader._save_data(cleaned_data, data_file, "data", cleaned=True)


if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")

    parser = argparse.ArgumentParser(description="LanheXLSXReader for LANHE Echem files", prog="LanheXLSXReader.py")
    parser.add_argument("path", type=str, help="Path to XLSX file or folder")
    parser.add_argument("-o", "--output", type=str, default="output", help="Output directory (default: output)")
    parser.add_argument("--no-clean", action="store_true", help="Also save original (non-cleaned) data")

    args = parser.parse_args()
    path = Path(args.path)

    if not path.exists():
        logger.error(f"Path not found: {path}")
        sys.exit(1)

    XLSXDataReader.save(path, args.output, save_cleaned=True, save_original=args.no_clean)
